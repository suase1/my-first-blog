from django.shortcuts import render, HttpResponse
from crawling.forms import *

from urllib.request import urlopen
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
import os
# Create your views here.

def main(request):
    return render(request, 'main.html')

def Crawling(request):

##---------------------< yachuk >---------------------##
##yachuk_get_Boots(mainpage)##
    mainpage = "http://ninetofive.cafe24.com/index.html"
    html = urlopen(mainpage)
    bs0bj = BeautifulSoup(html, "html.parser")

    for link in bs0bj.find("div",{"id":"container"}).find_all("a"):
        try:
            if "신발" in link.string:
                shoespage = ("http://ninetofive.cafe24.com"+link.attrs['href'])
        except TypeError: pass

##yachuk_check_page_number(page)##
    page_number=[]
    count=0
    html = urlopen(shoespage)
    bs0bj = BeautifulSoup(html, "html.parser")

    for link in bs0bj.find_all("a", {"title" : "페이지로 이동"}):
        count += 1
        page_number.append("http://ninetofive.cafe24.com/product/list.html" + link.attrs['href'])

##yachuk_get_CodeNumber(shoespage)##
    codeNumber=[]
    compare=re.compile("\([A-Z]*[0-9]+\)")

    i = 0
    while i <= (len(page_number)-1):
        html = urlopen(page_number[i])
        bs0bj = BeautifulSoup(html, "html.parser")
        i = i+1

        for code in bs0bj.find_all("a", {"class" : "name"}):
            if compare.search(str(code)):
                n = compare.search(str(code))
                m = n.group()
                codeNumber.append(m)
    yachuk_list = codeNumber

    print("---------------------yachuk CodeNumber---------------------")
    print(yachuk_list)

##---------------------< ssaka >---------------------##
##ssaka_check_page_number(page)##
    page = "http://www.ssaka.co.kr/product/pro_list?cate_re=10001&cate_re=10001&cate0=&cate1=1&order=item_serial&cate3=13"
    page_number_list=[]
    html = urlopen(page)
    bs0bj = BeautifulSoup(html, "html.parser")

    page_number_list.append(page)
    #첫 페이지 등록
    for link in bs0bj.find("div", {"class" : "paging"}).find_all("a"):
        try:
            if int(link.string) in range(0, 20):
                if int(link.string) == 1:
                    pass
                    #first page type : javascript for unmaking hyperlink
                else:
                    page_number_list.append("http://www.ssaka.co.kr/product/pro_list" + link.attrs['href'])

        except ValueError:
            pass

    #return page_number_list


##ssaka_get_CodeNumber(brandpage, yachuk_list)##
    Accordance_CodeNumber = []
    Accordance_PR_link = []
    compare = re.compile("\([0-9]+\)")
    #page_number_list = ssaka_check_page_number(brandpage)
    for page in page_number_list:
        html = urlopen(page)
        bs0bj = BeautifulSoup(html, "html.parser")

        for code in bs0bj.find("div", {"id":"sProlistArea"}).find_all("a"):
            try:
                # code's second contents is <img>
                n = compare.search(code.contents[1].attrs['title'])
                ssaka_codenumber = n.group()
                if ssaka_codenumber in yachuk_list:
                    ssaka_link = "http://www.ssaka.co.kr"+code.attrs['href']
                    Accordance_CodeNumber.append(ssaka_codenumber)
                    Accordance_PR_link.append(ssaka_link)
            except AttributeError: pass

            except IndexError: break
            #a in div are included page <a>href, the number of that is 8.
    #return Accordance_PR_link,Accordance_CodeNumber


    accordance_link = Accordance_PR_link
    accordance_codenumber = Accordance_CodeNumber
##ssaka_get_Inventory(accordance_link, accordance_codenumber)##
    i = 0
    j = 2
    l = 2
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "제품코드"
    ws['B1'] = "SIZE"
    ws['C1'] = "보유"
    for page in accordance_link:
        html = urlopen(page)
        bs0bj = BeautifulSoup(html, "html.parser")
        compare = re.compile("[a-z]+")
        for size_info in bs0bj.find("table",{"class":"op_table"}).find_all("td",{"class":"op_t1"}):
            if str(size_info.string) == "SIZE":
                for sibling1 in size_info.next_siblings:
                    if sibling1 =="\n": pass
                    elif compare.search(sibling1.string): pass
                    else:
                        ws['A' + str(j)] = accordance_codenumber[i]
                        ws['B'+ str(j)] = sibling1.string
                        j += 1
            elif str(size_info.string) == "보유":
                for sibling2 in size_info.next_siblings:
                    if sibling2 == "\n":
                        if size_info.next_sibling.next_sibling == None:
                            ws['C' + str(l)] = "0"
                            l = l+1
                        pass
                    else:
                        ws['C'+str(l)] = sibling2.find("b").string.split("↑")[0]
                        l=l+1
            else:
                pass
        i += 1
    desktoppath = os.path.expanduser('~')
    wb.save(desktoppath + "\\Desktop\\inventory_information\\ssaka_Inventory.xlsx")
    #return 0
    print("------------------compare test---------------------")
    print("--------------yachuk vs ssaka--------------")
    print(accordance_codenumber)

##---------------------< kika >---------------------##
##kika_check_page_number(page)#
    page = "http://www.aaasports.co.kr/front/productlist.php?code=001000000000&brandcode=2"
    html = urlopen(page)
    bs0bj = BeautifulSoup(html, "html.parser")
    input_list = []
    page_number_list = []
    compare = re.compile("[0-9]+")

    listnum = bs0bj.find(attrs={"name":"listnum"})
    sort = bs0bj.find(attrs={"name":"sort"})
    block = bs0bj.find(attrs={"name":"block"})
    gotopage = bs0bj.find(attrs={"name":"gotopage"})

    input_list.append(listnum)
    input_list.append(sort)
    input_list.append(block)
    input_list.append(gotopage)
    n =''
    for input in input_list:
        m = '&'+input.attrs['name']+'='+input.attrs['value']
        n = n+m
    first_page = page+n
    page_format = first_page.split('&gotopage')[0]

    for number in bs0bj.find_all(attrs={"class":"prlist"}):
        page_number = number.string

    last_page= compare.search(page_number)
    last_page_number = last_page.group()

    for i in range(1, int(last_page_number)+1):
        gotopage = (page_format+"&gotopage="+"%d"%i)
        page_number_list.append(gotopage)

    #return page_number_list

##kika_get_CodeNumber(brandpage, yachuk_list)#
    compare = re.compile("\([A-Z]*[0-9]+\)")
    Accordance_CodeNumber = []
    Accordance_PR_link = []
    #page_number_list = kika_check_page_number(brandpage)
    for page in page_number_list:
        html = urlopen(page)
        bs0bj = BeautifulSoup(html, "html.parser")

        for tag in bs0bj.find_all(attrs={"class":"mainprname"}):
            n=compare.search(tag.string)
            m=n.group()
            if m in yachuk_list:
                Accordance_CodeNumber.append(m)

                pr_href = tag.parent.attrs['href']
                pr_link = pr_href.split('..')[1]
                result = "http://www.aaasports.co.kr" + pr_link
                Accordance_PR_link.append(result)
    #return Accordance_PR_link, Accordance_CodeNumber

    accordance_link = Accordance_PR_link
    accordance_number = Accordance_CodeNumber
##kika_get_Inventory(accordance_link, accordance_number)##
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "제품코드"
    ws['B1'] = "SIZE"
    ws['C1'] = "재고"
    i=0
    l=0
    j=1

    compare = re.compile("[0-9]+")
    for page in accordance_link:
        html = urlopen(page)
        bs0bj = BeautifulSoup(html, "html.parser")

        for link in bs0bj.find_all("b", limit=2):
            if link.string =="제품스펙":
                parent = link.parent.parent
                for tag in parent.next_siblings:
                    if tag.name == "tr":
                        for t in tag.find_all("td"):
                            if t.string:
                                l=l+1
                                if l%2 !=0:
                                    j= j +1
                                    ws['A' + str(j)] = accordance_number[i]
                                    ws['B' + str(j)] = compare.search(str(t.string)).group()
                                else:
                                    ws['C' + str(j)] = t.string.split("↑")[0]
                            else: pass
                    else: pass
            else: pass
        i += 1
    desktoppath = os.path.expanduser('~')
    wb.save(desktoppath + "\\Desktop\\inventory_information\\kika_Inventory.xlsx")
    #return 0
    print("--------------yachuk vs kika--------------")
    print(accordance_number)


##<------------------------ fifa ---------------------------->##
##fifa_check_page_number(page)##
    page = "http://fifas.co.kr/html/sub/main.php?&cate=022&align=&item_num=&brand=%BE%C6%B5%F0%B4%D9%BD%BA"
    page_number_list=[]
    for i in range(1,30):
        page.split('?')
        page_number_list.append(page.split('?')[0] + "?page=%d"%i + page.split('?')[1])
    #return page_number


##fifa_get_CodeNumber(brandpage, yachuk_list)##
    compare = re.compile("[A-Z]*[0-9]+")
    Accordance_CodeNumber = []
    Accordance_PR_link = []
    #page_number_list = fifa_check_page_number(brandpage)
    for page in page_number_list:
        html = urlopen(page)
        bs0bj = BeautifulSoup(html, "html.parser")
        if bs0bj.find("div", {"class": "obj"}):
            for link in bs0bj.find_all("div", {"class":"obj"}):
                n = compare.search(link.find('em').string)
                m = '('+n.group()+')'
                if m in yachuk_list:
                    Accordance_CodeNumber.append(m)
                    Accordance_PR_link.append("http://fifas.co.kr" + link.find("a").attrs['href'])
        else:
            break
            #so as to remove the empty page
    #return Accordance_PR_link, Accordance_CodeNumber

    accordance_link = Accordance_PR_link
    accordance_number = Accordance_CodeNumber

##fifa_get_Inventory(accordance_link, accordance_number)##
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "제품코드"
    ws['B1'] = "SIZE"
    ws['C1'] = "재고"
    i = 0
    l = 0
    j = 1
    for page in accordance_link:
        html = urlopen(page)
        bs0bj = BeautifulSoup(html, "html.parser")
        compare = re.compile("[0-9]+")
        for link in bs0bj.find("td",{"class":"s_text01"}).parent.next_siblings:
            try:
                for tag in link.children:
                    if tag == "\n": pass
                    elif compare.match(str(tag.contents[0])):
                            l = l+1
                            if l%2 != 0:
                                j = j +1
                                ws['A' + str(j)] = accordance_number[i]
                                ws['B' + str(j)] = tag.contents[0]
                            else:
                                ws['C'+str(j)] = tag.contents[0].split("↑")[0]
                    else: pass
            except AttributeError: pass
        i += 1
    desktoppath = os.path.expanduser('~')
    wb.save(desktoppath + "\\Desktop\\inventory_information\\fifa_Inventory.xlsx")
    #return 0
    print("--------------yachuk vs fifa--------------")
    print(accordance_number)

##<----------------Integration_version(yachuk_excel, ssaka_excel, fifa_excel, kika_excel)-------------->##
    yachuk_excel = desktoppath + "\\Desktop\\inventory_information\\ninetofive.xlsx"
    ssaka_excel = desktoppath + "\\Desktop\\inventory_information\\ssaka_Inventory.xlsx"
    fifa_excel = desktoppath + "\\Desktop\\inventory_information\\fifa_Inventory.xlsx"
    kika_excel = desktoppath + "\\Desktop\\inventory_information\\kika_Inventory.xlsx"

    wb_Yachuk = load_workbook(yachuk_excel)
    wb_ssaka = load_workbook(ssaka_excel)
    wb_fifa = load_workbook(fifa_excel)
    wb_kika = load_workbook(kika_excel)

    ws_Yachuk = wb_Yachuk.active
    ws_ssaka = wb_ssaka.active
    ws_fifa = wb_fifa.active
    ws_kika = wb_kika.active

    for i in range(2,len(list(ws_Yachuk.rows))+1):
        ws_Yachuk.cell(row=i, column=10).value = 0

    for j in range(2, len(list(ws_ssaka.rows))+1):
        for l in range(2,len(list(ws_Yachuk.rows))+1):
            if str(ws_ssaka.cell(row =j, column=1).value) in str(ws_Yachuk.cell(row=l, column=3).value) and str(ws_ssaka.cell(row=j, column=2).value) == str(ws_Yachuk.cell(row=l, column=7).value):
                ws_Yachuk.cell(row=l, column=10).value = ws_ssaka.cell(row=j, column=3).value
                break
            else: pass
    j = 0
    l = 0
    for j in range(2, len(list(ws_fifa.rows))+1):
        for l in range(2,len(list(ws_Yachuk.rows))+1):
            if str(ws_fifa.cell(row =j, column=1).value) in str(ws_Yachuk.cell(row=l, column=3).value) and str(ws_fifa.cell(row=j, column=2).value) == str(ws_Yachuk.cell(row=l, column=7).value):
                ws_Yachuk.cell(row=l, column=10).value = ws_fifa.cell(row=j, column=3).value
                break
            else: pass
    j = 0
    l = 0
    for j in range(2, len(list(ws_kika.rows))+1):
        for l in range(2,len(list(ws_Yachuk.rows))+1):
            if str(ws_kika.cell(row =j, column=1).value) in str(ws_Yachuk.cell(row=l, column=3).value) and str(ws_kika.cell(row=j, column=2).value) == str(ws_Yachuk.cell(row=l, column=7).value):
                ws_Yachuk.cell(row=l, column=10).value = int(ws_Yachuk.cell(row=l, column=10).value) + int(ws_kika.cell(row=j, column=3).value)
                break
            else: pass

    desktoppath = os.path.expanduser('~')
    wb_Yachuk.save(desktoppath+"\\Desktop\\inventory_information\\Integration_version.xlsx")
    return HttpResponse("FINISH")
